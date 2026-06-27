"""
app/routes/reports.py
======================
Static-asset and reporting routes.

GET /logo.png       — app icon (PWA)
GET /manifest.json  — PWA manifest
GET /rules          — Rules.md as JSON for frontend markdown rendering
GET /export_xlsx    — Full drink-log XLSX download for the session
GET /summary_json   — Drink summary as JSON for on-screen display
"""

import csv
import io
import os
from collections import defaultdict
from datetime import datetime
from zoneinfo import ZoneInfo

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from flask import Blueprint, current_app, jsonify, request, Response, send_from_directory

from app.services.session_store import game_sessions

bp = Blueprint("reports", __name__)

# ── Project root (one level above the app/ package) ─────────────────────────
_HERE = os.path.dirname(os.path.abspath(__file__))   # .../app/routes/
_ROOT = os.path.dirname(os.path.dirname(_HERE))      # .../Black-Out-Jack/


# ---------------------------------------------------------------------------
# Static assets
# ---------------------------------------------------------------------------

@bp.route("/logo.png")
def serve_logo():
    return send_from_directory(current_app.static_folder, "img/logo.png")


@bp.route("/manifest.json")
def serve_manifest():
    return jsonify({
        "name":             "Black(Out)Jack",
        "short_name":       "Black(Out)Jack",
        "start_url":        "/",
        "display":          "standalone",
        "background_color": "#0f1117",
        "theme_color":      "#0f1117",
        "icons": [
            {"src": "/logo.png", "sizes": "192x192", "type": "image/png"},
            {"src": "/logo.png", "sizes": "512x512", "type": "image/png"},
        ],
    })


# ---------------------------------------------------------------------------
# Rules
# ---------------------------------------------------------------------------

@bp.route("/rules")
def rules():
    """Serve the Rules.md content as plain text for frontend markdown rendering."""
    rules_path = os.path.join(_ROOT, "docs", "Rules.md")
    try:
        with open(rules_path, "r", encoding="utf-8") as f:
            content = f.read()
        return jsonify({"ok": True, "content": content})
    except FileNotFoundError:
        return jsonify({"ok": False, "content": "# Rules\n\nRules file not found."})


# ---------------------------------------------------------------------------
# Styling helpers for XLSX export
# ---------------------------------------------------------------------------

_FONT_BOLD    = Font(bold=True)
_FONT_SECTION = Font(bold=True, color="FFFFFF")
_FILL_SECTION = PatternFill("solid", fgColor="1F4E79")   # dark navy
_FILL_HEADER  = PatternFill("solid", fgColor="BDD7EE")   # light blue


def _xlsx_section(ws, title: str) -> None:
    """Append a dark-navy section-title row."""
    ws.append([title])
    cell = ws.cell(row=ws.max_row, column=1)
    cell.font = _FONT_SECTION
    cell.fill = _FILL_SECTION


def _xlsx_header(ws, cols: list) -> None:
    """Append a light-blue column-header row."""
    ws.append(cols)
    r = ws.max_row
    for c in range(1, len(cols) + 1):
        cell = ws.cell(row=r, column=c)
        cell.font = _FONT_BOLD
        cell.fill = _FILL_HEADER


def _auto_width(ws) -> None:
    """Set column widths based on content length (capped at 40)."""
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 40)


# ---------------------------------------------------------------------------
# XLSX export
# ---------------------------------------------------------------------------

@bp.route("/export_xlsx")
def export_xlsx():
    """
    Return an Excel workbook summarising drinks recorded in this session.
    Sheet 1 "Summary"  — formatted session report (metadata, per-player tables, totals).
    Sheet 2 "Raw Rows" — flat drink log, one row per drink event.
    Usage: GET /export_xlsx?room_code=Jack-21
    """
    room_code = request.args.get("room_code", "")
    session   = game_sessions.get(room_code)
    if not session:
        return Response("No active session.", status=404, mimetype="text/plain")

    rows         = session.drinks.csv_rows
    hand_stats   = session.stats.hand_stats
    milestones   = session.drinks.milestones_claimed
    dealer_stats = session.stats.dealer_hand_stats
    wc_presses   = session.drinks.wild_card_presses

    num_rounds = max((r["round"] for r in rows), default=1)
    players_seen: list[str] = []
    player_sips: dict[str, dict[str, int]] = defaultdict(lambda: defaultdict(int))
    dealer_sips: dict[str, dict[str, int]] = defaultdict(lambda: defaultdict(int))

    for row in rows:
        name = row["player"]
        if name not in players_seen:
            players_seen.append(name)
        bucket = dealer_sips if row["role"] == "dealer" else player_sips
        bucket[name][row["rule"]] += row["sips"]

    all_rules = sorted({row["rule"] for row in rows})

    def _pct(n, d):
        return f"{n/d*100:.1f}%" if d else "—"

    _tz = ZoneInfo("Europe/Zurich")

    wb = openpyxl.Workbook()

    # ── Sheet 1: Summary ─────────────────────────────────────────────
    ws = wb.active
    ws.title = "Summary"

    # Metadata block
    _xlsx_section(ws, "Drinking Blackjack — Session Summary")
    ws.append(["Generated",        datetime.now(_tz).strftime("%Y-%m-%d %H:%M:%S")])
    ws.append(["Rounds completed", num_rounds])
    ws.append(["Players"] + players_seen)
    ws.append([])

    # Milestones
    if milestones:
        _xlsx_section(ws, "MILESTONES")
        _xlsx_header(ws, ["Threshold", "First to reach"])
        for boundary in sorted(milestones):
            ws.append([f"{boundary} sips", milestones[boundary]])
        ws.append([])

    # Wild Card
    if wc_presses:
        _xlsx_section(ws, "WILD CARD 🃏")
        _xlsx_header(ws, ["Player", "Total presses", "Self", "Random", "Dud"])
        for name, s in sorted(wc_presses.items(), key=lambda x: -x[1]["presses"]):
            ws.append([name, s["presses"], s["self"], s["random"], s["dud"]])
        ws.append([])

    # Per-player sections
    for name in players_seen:
        pt = sum(player_sips[name].values())
        dt = sum(dealer_sips[name].values())
        gt = pt + dt
        hs = hand_stats.get(name)
        h  = hs["hands"] if hs else 0

        _xlsx_section(ws, name)
        ws.append([f"Total sips: {gt}", f"As player: {pt}", f"As dealer: {dt}",
                   f"Sips/round: {gt/num_rounds:.2f}", f"Hands: {h}" if h else "Hands: 0"])
        if hs and h:
            ws.append([
                f"Won: {hs['wins']} ({_pct(hs['wins'], h)})",
                f"Lost: {hs['losses']} ({_pct(hs['losses'], h)})",
                f"Push: {hs['pushes']} ({_pct(hs['pushes'], h)})",
                (f"Splits won: {hs['split_wins']} of {hs['split_hands']}"
                 f" ({_pct(hs['split_wins'], hs['split_hands'])})" if hs["split_hands"] else ""),
                (f"Doubles won: {hs['double_wins']} of {hs['double_hands']}"
                 f" ({_pct(hs['double_wins'], hs['double_hands'])})" if hs["double_hands"] else ""),
            ])
        _xlsx_header(ws, ["Rule", "Player sips", "Dealer sips", "Total", "Sips/round", "% of own"])
        player_rules = [
            (rule, player_sips[name].get(rule, 0) + dealer_sips[name].get(rule, 0))
            for rule in all_rules
        ]
        for rule, total in sorted(player_rules, key=lambda x: -x[1]):
            if total == 0:
                continue
            ps  = player_sips[name].get(rule, 0)
            ds  = dealer_sips[name].get(rule, 0)
            pct = f"{total/gt*100:.1f}%" if gt else "—"
            ws.append([rule, ps, ds, total, f"{total/num_rounds:.2f}", pct])
        ws.append(["TOTAL", pt, dt, gt, f"{gt/num_rounds:.2f}", "100%"])
        ws.cell(row=ws.max_row, column=1).font = _FONT_BOLD
        ws.append([])

    # Grand totals
    rule_totals: dict[str, int] = defaultdict(int)
    for name in players_seen:
        for rule, s in player_sips[name].items():
            rule_totals[rule] += s
        for rule, s in dealer_sips[name].items():
            rule_totals[rule] += s
    grand_total = sum(rule_totals.values())

    _xlsx_section(ws, "ALL PLAYERS COMBINED")
    _xlsx_header(ws, ["Rule", "Total sips", "Sips/round", "% of total"])
    for rule in sorted(rule_totals, key=lambda r: -rule_totals[r]):
        total = rule_totals[rule]
        pct   = f"{total/grand_total*100:.1f}%" if grand_total else "—"
        ws.append([rule, total, f"{total/num_rounds:.2f}", pct])
    ws.append(["Grand total", grand_total, f"{grand_total/num_rounds:.2f} sips/round"])
    ws.cell(row=ws.max_row, column=1).font = _FONT_BOLD
    ws.append([])

    # Hand outcomes
    _xlsx_section(ws, "HAND OUTCOMES")
    _xlsx_header(ws, ["Player", "Hands", "Won", "Win%", "Lost", "Loss%", "Push", "Push%",
                      "Splits won", "Split win%", "Doubles won", "Double win%"])
    for name in players_seen:
        hs = hand_stats.get(name, {
            "hands": 0, "wins": 0, "losses": 0, "pushes": 0,
            "split_hands": 0, "split_wins": 0, "double_hands": 0, "double_wins": 0,
        })
        h = hs["hands"]
        ws.append([
            name, h if h else "-",
            hs["wins"]   if h else "-", _pct(hs["wins"],   h),
            hs["losses"] if h else "-", _pct(hs["losses"],  h),
            hs["pushes"] if h else "-", _pct(hs["pushes"],  h),
            f"{hs['split_wins']} of {hs['split_hands']}" if hs["split_hands"]  else "-",
            _pct(hs["split_wins"],  hs["split_hands"]),
            f"{hs['double_wins']} of {hs['double_hands']}" if hs["double_hands"] else "-",
            _pct(hs["double_wins"], hs["double_hands"]),
        ])

    # Dealer stats
    if dealer_stats:
        ws.append([])
        _xlsx_section(ws, "DEALER STATS (per dealing stint)")
        _xlsx_header(ws, ["Dealer", "Hands dealt", "Won", "Win%", "Lost", "Loss%", "Push", "Push%"])
        for dname, ds in sorted(dealer_stats.items()):
            dh = ds["hands"]
            ws.append([
                dname, dh,
                ds["wins"],   _pct(ds["wins"],   dh),
                ds["losses"], _pct(ds["losses"],  dh),
                ds["pushes"], _pct(ds["pushes"],  dh),
            ])

    _auto_width(ws)

    # ── Sheet 2: Raw Rows ────────────────────────────────────────
    ws2 = wb.create_sheet("Raw Rows")
    _xlsx_header(ws2, ["round", "dealer", "player", "role", "rule", "sips"])
    for row in rows:
        ws2.append([row.get("round"), row.get("dealer"), row.get("player"),
                    row.get("role"), row.get("rule"), row.get("sips")])
    _auto_width(ws2)

    # ── Respond ────────────────────────────────────────────────────────────────
    _date_str = datetime.now(_tz).strftime("%Y-%m-%d")
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        buf.getvalue(),
        status=200,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="drinks_summary_{_date_str}.xlsx"'},
    )


# ---------------------------------------------------------------------------
# Decision log export (Phase C — see docs/planning/DecisionLog-Plan.md)
# ---------------------------------------------------------------------------

_DECISION_COLUMNS = [
    "session_id", "timestamp", "round", "player", "hand_index",
    "dealer_name", "hand_cards_before", "hand_total_before", "is_soft",
    "dealer_upcard", "visible_cards", "cards_remaining", "decks_in_play",
    "valid_actions", "action_taken", "basic_strategy_action",
    "drinking_mode", "mode", "bet_amount", "wager", "is_npc", "hand_result",
]


@bp.route("/export_decisions")
def export_decisions():
    """
    Return a CSV of recorded player decisions for this session.
    Usage: GET /export_decisions?room_code=Jack-21&player=<name optional>
    """
    room_code = request.args.get("room_code", "")
    player    = request.args.get("player", "").strip()

    session = game_sessions.get(room_code)
    if not session:
        return Response("No active session.", status=404, mimetype="text/plain")

    rows = session._decision_log
    if player:
        rows = [r for r in rows if r["player"].lower() == player.lower()]

    buf = io.StringIO()
    w   = csv.writer(buf)
    w.writerow(_DECISION_COLUMNS)
    for row in rows:
        w.writerow([row.get(col, "") for col in _DECISION_COLUMNS])

    _date_str = datetime.now(ZoneInfo("Europe/Zurich")).strftime("%Y-%m-%d")
    return Response(
        b"\xef\xbb\xbf" + buf.getvalue().encode("utf-8"),  # UTF-8 BOM for Excel
        status=200,
        mimetype="text/csv",
        headers={"Content-Disposition": f'attachment; filename="decision_log_{_date_str}.csv"'},
    )


# ---------------------------------------------------------------------------
# JSON summary
# ---------------------------------------------------------------------------

@bp.route("/summary_json")
def summary_json():
    """Return session drink summary as JSON for on-screen display."""
    room_code = request.args.get("room_code", "")
    session   = game_sessions.get(room_code)
    if not session:
        return jsonify({"ok": False, "error": "Room not found."})

    rows       = session.drinks.csv_rows
    num_rounds = max((r["round"] for r in rows), default=0)

    player_sips: dict[str, int] = defaultdict(int)
    dealer_sips: dict[str, int] = defaultdict(int)
    players_seen: list[str]     = []

    for row in rows:
        name = row["player"]
        if name not in players_seen:
            players_seen.append(name)
        if row["role"] == "dealer":
            dealer_sips[name] += row["sips"]
        else:
            player_sips[name] += row["sips"]

    summary = []
    for name in players_seen:
        ps = player_sips[name]
        ds = dealer_sips[name]
        summary.append({"name": name, "player_sips": ps,
                         "dealer_sips": ds, "total_sips": ps + ds})
    summary.sort(key=lambda x: -x["total_sips"])

    return jsonify({"ok": True, "rounds": num_rounds, "players": summary})
