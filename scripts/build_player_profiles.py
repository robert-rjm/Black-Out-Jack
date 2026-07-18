"""
scripts/build_player_profiles.py
==================================
Build per-player deviation tables from accumulated decision-log workbooks and
write them to engine/player_profiles/<name>.json.

Each decision_log_*.xlsx (downloaded from GET /export_decisions and saved
into data/decisions/) has two sheets: "Hand Decisions" (hit/stand/double/
split/insurance) and "Dealer Lottery Entries" (the 0-5 stake decision) --
both get mined here into the same profile JSON, under "deviations" and
"lottery_stakes" respectively. Older decision_log_*.csv exports (single
"Hand Decisions" sheet, from before the two-sheet xlsx format existed) are
also read -- same columns, no Dealer Lottery data.

Usage:
    python scripts/build_player_profiles.py [--dir data/decisions] [--out engine/player_profiles]
    python scripts/build_player_profiles.py --player Rob  # one player only
    python scripts/build_player_profiles.py --merge       # fold in the existing
                                                            # profile's own recorded
                                                            # deviations instead of
                                                            # overwriting from scratch

Thresholds (adjustable via flags):
    --min-samples   Minimum decisions at a spot to record anything  (default: 3)
    --min-majority  Minimum fraction for the majority action         (default: 0.60)

Only deviations from basic_strategy_action are stored; spots where the player
agrees with basic strategy are omitted (the fallback handles those).

--merge: a profile JSON only ever records spots that already qualified as a
deviation -- it has no memory of spots where the player agreed with basic
strategy, so this can only ever recover *some* of what a from-scratch build
over the original raw logs would have shown, but it exactly preserves every
recorded action_counts and folds them in with whatever the current raw logs
produce, so a spot's sample size/majority reflects the true combined history
rather than being reset to whatever's newly available.
"""

from __future__ import annotations

import argparse
import csv
import json
import os
import sys
from collections import defaultdict
from datetime import date
from pathlib import Path

import openpyxl

# Repo root on sys.path so we can reuse the table_bias/owed-sips bucket
# thresholds from engine/style_strategy.py instead of re-deriving them here.
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from engine.style_strategy import _table_bias_bucket as _bias_bucket  # noqa: E402
from engine.style_strategy import _owed_bucket  # noqa: E402


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _rank(card_str: str) -> str:
    """Extract rank from a card string like '7♣', 'J♦', '10♥', 'A♠'."""
    # Suits are single unicode chars; strip the last character.
    return card_str[:-1] if card_str else ""


def _hand_index(row: dict) -> int:
    try:
        return int(row.get("hand_index") or 0)
    except ValueError:
        return 0


def _majority(actions: list[str]) -> tuple[str, float, dict[str, int]]:
    """Return (majority_action, majority_fraction, action_counts) for a list
    of action strings."""
    counts: dict[str, int] = defaultdict(int)
    for a in actions:
        counts[a] += 1
    majority_action = max(counts, key=lambda a: counts[a])
    return majority_action, counts[majority_action] / len(actions), dict(counts)


def _load_xlsx_sheet(directory: str, sheet_name: str,
                      pattern: str = "decision_log_*.xlsx") -> list[dict]:
    """
    Load every row of `sheet_name` across every decision_log_*.xlsx file in
    `directory` (each workbook has a "Hand Decisions" sheet and a
    "Dealer Lottery Entries" sheet -- see /export_decisions in
    app/routes/reports.py). Values are coerced to strings so the result
    matches csv.DictReader's shape exactly, since every downstream parser
    here (_parse_bool, int(), .strip(), etc.) was written against that
    all-strings assumption and is unchanged by this xlsx switch.
    """
    rows = []
    d = Path(directory)
    for path in sorted(d.glob(pattern)):
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        row_iter = ws.iter_rows(values_only=True)
        header = next(row_iter, None)
        if not header:
            continue
        for values in row_iter:
            rows.append({
                str(h): ("" if v is None else str(v))
                for h, v in zip(header, values)
            })
        print(f"  loaded {path.name} [{sheet_name}]")
    return rows


def _load_csv_sheet(directory: str, pattern: str = "decision_log_*.csv") -> list[dict]:
    """
    Load every row across every decision_log_*.csv file in `directory` --
    the older single-sheet "Hand Decisions" export format, from before
    /export_decisions became a two-sheet xlsx workbook (see reports.py's
    own docstring: "was /export_decisions as a standalone CSV"). Same
    columns as the "Hand Decisions" xlsx sheet, so the returned rows are
    interchangeable with _load_xlsx_sheet's. There's no CSV equivalent of
    the "Dealer Lottery Entries" sheet -- that was a separate export this
    old format never had, so lottery_stakes simply can't be recovered from
    these files.
    """
    rows = []
    d = Path(directory)
    for path in sorted(d.glob(pattern)):
        # utf-8-sig strips a leading BOM if present (Excel/PowerShell-style
        # CSV exports commonly have one) so the first header key isn't
        # mangled into "﻿session_id".
        with open(path, "r", encoding="utf-8-sig", newline="") as f:
            for row in csv.DictReader(f):
                rows.append({k: (v if v is not None else "") for k, v in row.items()})
        print(f"  loaded {path.name} [csv]")
    return rows


def _parse_bool(val: str) -> bool:
    return val.strip().lower() == "true"


# ---------------------------------------------------------------------------
# Merge support -- fold an existing profile's own recorded deviations back
# in as if they were extra raw rows (see --merge in the module docstring)
# ---------------------------------------------------------------------------

def _counts_from_deviations(deviations: list[dict]) -> tuple[dict, dict, dict, dict]:
    """
    Reconstruct per-key action counts from a profile's already-recorded
    deviations -- the only place old per-spot data survives once the raw
    decision_log_*.xlsx/csv files that produced them are gone. Returns
    (coarse_counts, coarse_basic, fine_counts, fine_basic), the same shape
    build_profile builds internally from raw rows, so the two can be
    combined by simple dict-of-dict addition.

    Only spots that were ALREADY a deviation carry any historical weight
    this way -- a spot where the player agreed with basic strategy (or
    that never reached min_samples/min_majority) was never recorded in the
    first place and can't be recovered from the profile alone.
    """
    coarse_counts: dict[tuple, dict[str, int]] = defaultdict(lambda: defaultdict(int))
    coarse_basic: dict[tuple, str] = {}
    fine_counts: dict[tuple, dict[str, int]] = defaultdict(lambda: defaultdict(int))
    fine_basic: dict[tuple, str] = {}

    for d in deviations:
        base_key = (d["hand_total"], d["is_soft"], d["dealer_upcard_rank"],
                    d["can_split"], d["can_double"])
        if "table_bias" in d:
            fine_key = base_key + (d["table_bias"], d["sibling_awaiting_deal"])
            for action, n in d["action_counts"].items():
                fine_counts[fine_key][action] += n
            fine_basic[fine_key] = d["basic_strategy"]
        else:
            for action, n in d["action_counts"].items():
                coarse_counts[base_key][action] += n
            coarse_basic[base_key] = d["basic_strategy"]

    return dict(coarse_counts), coarse_basic, dict(fine_counts), fine_basic


def merge_lottery_stakes(old_stakes: list[dict], new_stakes: list[dict]) -> list[dict]:
    """
    Combine two lottery_stakes lists (each {owed_bucket, avg_stake, samples})
    per bucket, reconstructing each side's sum from avg_stake * samples
    (exact, modulo the 2-decimal rounding already baked into avg_stake) so
    the merged average is a true weighted combination rather than a naive
    average-of-averages.
    """
    totals: dict[str, dict[str, float]] = defaultdict(lambda: {"sum": 0.0, "samples": 0})
    for s in (old_stakes or []) + (new_stakes or []):
        t = totals[s["owed_bucket"]]
        t["sum"]     += s["avg_stake"] * s["samples"]
        t["samples"] += s["samples"]

    merged = [
        {
            "owed_bucket": bucket,
            "avg_stake":   round(t["sum"] / t["samples"], 2),
            "samples":     t["samples"],
        }
        for bucket, t in totals.items() if t["samples"] > 0
    ]
    merged.sort(key=lambda s: s["owed_bucket"])
    return merged


# ---------------------------------------------------------------------------
# Core builder
# ---------------------------------------------------------------------------

def build_profile(rows: list[dict], player_name: str,
                  min_samples: int = 3, min_majority: float = 0.60,
                  extra_counts: tuple | None = None,
                  extra_source_decisions: int = 0) -> dict:
    """
    Return a profile dict for one player.

    Groups decisions by (hand_total, is_soft, dealer_upcard_rank, can_split,
    can_double) and records a deviation entry where the player's majority
    action differs from basic_strategy_action and meets the thresholds.

    Also groups the same decisions by that same key *plus* table_bias
    (bucketed from the row's visible_cards) and sibling_awaiting_deal
    (whether another of the player's hands this round, from a split, was
    still undealt at decision time). A table-aware deviation is only kept
    when its majority action differs from what the coarser (plain 5-field)
    grouping would already produce -- otherwise the extra context is noise
    and the coarser deviation (or basic strategy) already covers the spot.

    `extra_counts` (see --merge / _counts_from_deviations) is folded in by
    extending each key's action list with the extra counts BEFORE any
    majority/threshold logic runs -- extending a key's list with N copies
    of an action is exactly equivalent to having N more raw rows at that
    spot, so every existing computation below runs unmodified over the
    combined history. `extra_source_decisions` is just added to the
    reported `source_decisions` count for an honest cumulative total.
    """
    # Filter: this player, human only, non-insurance, has a basic_strategy_action
    player_rows = [
        r for r in rows
        if r["player"].strip().lower() == player_name.lower()
        and not _parse_bool(r.get("is_npc", "False"))
        and r.get("action_taken", "") not in ("insurance", "decline")
        and r.get("basic_strategy_action", "").strip()
    ]

    # A hand_index greater than this row's own, anywhere in the same round
    # for this player, means that hand was still undealt (1 card, from a
    # split) at the moment of this decision -- splits are dealt and played
    # strictly in hand_index order (see engine/style_strategy.py's
    # _sibling_awaiting_deal docstring for why).
    max_hand_index_by_round: dict[str, int] = defaultdict(int)
    for r in player_rows:
        rnd = r.get("round", "")
        max_hand_index_by_round[rnd] = max(max_hand_index_by_round[rnd], _hand_index(r))

    groups: dict[tuple, list[str]] = defaultdict(list)
    basic_for_key: dict[tuple, str] = {}
    fine_groups: dict[tuple, list[str]] = defaultdict(list)
    fine_basic_for_key: dict[tuple, str] = {}

    for r in player_rows:
        valid = r.get("valid_actions", "")
        can_split  = "sp" in valid
        can_double = "d"  in valid
        try:
            total = int(r["hand_total_before"])
        except (ValueError, KeyError):
            continue
        is_soft     = _parse_bool(r.get("is_soft", "False"))
        dealer_rank = _rank(r.get("dealer_upcard", "").strip())
        if not dealer_rank:
            continue
        action    = r["action_taken"].strip()
        bs_action = r["basic_strategy_action"].strip()

        base_key = (total, is_soft, dealer_rank, can_split, can_double)
        groups[base_key].append(action)
        basic_for_key[base_key] = bs_action  # same for any row at this key

        bias    = _bias_bucket(r.get("visible_cards", "").split())
        sibling = max_hand_index_by_round[r.get("round", "")] > _hand_index(r)
        fine_key = base_key + (bias, sibling)
        fine_groups[fine_key].append(action)
        fine_basic_for_key[fine_key] = bs_action

    if extra_counts:
        extra_coarse, extra_coarse_basic, extra_fine, extra_fine_basic = extra_counts
        for key, counts in extra_coarse.items():
            for action, n in counts.items():
                groups[key].extend([action] * n)
            basic_for_key.setdefault(key, extra_coarse_basic[key])
        for key, counts in extra_fine.items():
            for action, n in counts.items():
                fine_groups[key].extend([action] * n)
            fine_basic_for_key.setdefault(key, extra_fine_basic[key])

    # --- Coarse (5-field) deviations -- unchanged behavior ---
    deviations = []
    coarse_action_by_key: dict[tuple, str] = {}
    for key, actions in groups.items():
        n = len(actions)
        if n < min_samples:
            continue
        majority_action, majority_frac, counts = _majority(actions)
        if majority_frac < min_majority:
            continue
        bs = basic_for_key[key]
        if majority_action == bs:
            continue  # agrees with basic strategy — no deviation to store

        coarse_action_by_key[key] = majority_action
        total, is_soft, dealer_rank, can_split, can_double = key
        deviations.append({
            "hand_total":        total,
            "is_soft":           is_soft,
            "dealer_upcard_rank": dealer_rank,
            "can_split":         can_split,
            "can_double":        can_double,
            "basic_strategy":    bs,
            "player_action":     majority_action,
            "samples":           n,
            "majority":          round(majority_frac, 3),
            "action_counts":     counts,
        })

    # --- Fine (table-aware) deviations -- only where the extra context
    # actually reveals a different tendency than the coarse tier above ---
    for fine_key, actions in fine_groups.items():
        n = len(actions)
        if n < min_samples:
            continue
        majority_action, majority_frac, counts = _majority(actions)
        if majority_frac < min_majority:
            continue

        base_key = fine_key[:5]
        bias, sibling = fine_key[5], fine_key[6]
        bs = fine_basic_for_key[fine_key]
        baseline = coarse_action_by_key.get(base_key, bs)
        if majority_action == baseline:
            continue  # no different from the coarser spot -- not worth recording

        total, is_soft, dealer_rank, can_split, can_double = base_key
        deviations.append({
            "hand_total":            total,
            "is_soft":               is_soft,
            "dealer_upcard_rank":    dealer_rank,
            "can_split":             can_split,
            "can_double":            can_double,
            "table_bias":            bias,
            "sibling_awaiting_deal": sibling,
            "basic_strategy":        bs,
            "player_action":         majority_action,
            "samples":               n,
            "majority":              round(majority_frac, 3),
            "action_counts":         counts,
        })

    deviations.sort(key=lambda d: (d["hand_total"], d["is_soft"], d["dealer_upcard_rank"],
                                   d.get("table_bias") or "", d.get("sibling_awaiting_deal", False)))

    return {
        "player":           player_name,
        "generated":        str(date.today()),
        "source_decisions": len(player_rows) + extra_source_decisions,
        "thresholds": {
            "min_samples":  min_samples,
            "min_majority": min_majority,
        },
        "deviations": deviations,
    }


def build_lottery_stakes(rows: list[dict], player_name: str,
                          min_samples: int = 3) -> list[dict]:
    """
    Group this player's (human) Dealer Lottery entries -- rows from the
    "Dealer Lottery Entries" sheet of decision_log_*.xlsx, written by
    app.services.decision_log.record_dealer_lottery_entry -- by how many
    sips they owed this round when the entry window opened, and record
    their average stake per bucket.

    This is what engine/style_strategy.py's decide_dealer_lottery_stake
    looks up to drive an NPC using this profile; a bucket without enough
    samples is simply omitted (NPC falls back to opting out, same as a
    "basic" personality).
    """
    player_rows = [
        r for r in rows
        if r["player"].strip().lower() == player_name.lower()
        and not _parse_bool(r.get("is_npc", "False"))
    ]

    buckets: dict[str, list[int]] = defaultdict(list)
    for r in player_rows:
        try:
            owed = int(r.get("current_owed", "") or 0)
            x    = int(r.get("x_entered", "") or 0)
        except ValueError:
            continue
        buckets[_owed_bucket(owed)].append(x)

    stakes = []
    for bucket, values in buckets.items():
        n = len(values)
        if n < min_samples:
            continue
        stakes.append({
            "owed_bucket": bucket,
            "avg_stake":   round(sum(values) / n, 2),
            "samples":     n,
        })
    stakes.sort(key=lambda s: s["owed_bucket"])
    return stakes


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="Build player deviation profiles.")
    parser.add_argument("--dir",          default="data/decisions",
                        help="Directory containing decision_log_*.xlsx/.csv files "
                             "(downloaded from /export_decisions)")
    parser.add_argument("--out",          default="engine/player_profiles",
                        help="Output directory for <name>.json profiles")
    parser.add_argument("--player",       default=None,
                        help="Build profile for one player only")
    parser.add_argument("--min-samples",  type=int,   default=3)
    parser.add_argument("--min-majority", type=float, default=0.60)
    parser.add_argument("--merge", action="store_true",
                        help="Fold each player's EXISTING profile (if any) in "
                             "as extra history instead of overwriting from "
                             "scratch -- see the --merge note in this file's "
                             "module docstring for what it can and can't recover.")
    args = parser.parse_args()

    print(f"Loading decision logs from {args.dir} ...")
    rows = _load_xlsx_sheet(args.dir, "Hand Decisions") + _load_csv_sheet(args.dir)
    if not rows:
        print("No rows loaded — check --dir path.")
        sys.exit(1)

    lottery_rows = _load_xlsx_sheet(args.dir, "Dealer Lottery Entries")

    # Discover players
    all_players = sorted({r["player"].strip() for r in rows
                          if not _parse_bool(r.get("is_npc", "False"))})
    targets = [args.player] if args.player else all_players

    os.makedirs(args.out, exist_ok=True)

    for name in targets:
        out_path = Path(args.out) / f"{name.lower()}.json"

        extra_counts = None
        extra_source_decisions = 0
        old_lottery_stakes = []
        if args.merge and out_path.exists():
            with open(out_path, encoding="utf-8") as f:
                old_profile = json.load(f)
            extra_counts = _counts_from_deviations(old_profile.get("deviations", []))
            extra_source_decisions = old_profile.get("source_decisions", 0)
            old_lottery_stakes = old_profile.get("lottery_stakes", [])
            print(f"  merging with existing {out_path.name} "
                  f"({extra_source_decisions} prior decisions, "
                  f"{len(old_profile.get('deviations', []))} recorded deviation(s))")

        profile = build_profile(rows, name, args.min_samples, args.min_majority,
                                 extra_counts=extra_counts,
                                 extra_source_decisions=extra_source_decisions)
        new_lottery_stakes = build_lottery_stakes(lottery_rows, name, args.min_samples)
        profile["lottery_stakes"] = (
            merge_lottery_stakes(old_lottery_stakes, new_lottery_stakes)
            if args.merge else new_lottery_stakes
        )
        n_dev   = len(profile["deviations"])
        n_stake = len(profile["lottery_stakes"])
        with open(out_path, "w", encoding="utf-8") as f:
            json.dump(profile, f, indent=2)
        print(f"\n{name}: {profile['source_decisions']} decisions, "
              f"{n_dev} deviation(s), {n_stake} lottery-stake bucket(s) recorded → {out_path}")
        for d in profile["deviations"]:
            hand_desc = f"{'soft' if d['is_soft'] else 'hard'} {d['hand_total']}"
            flags = []
            if d["can_split"]:  flags.append("splittable")
            if d["can_double"]: flags.append("doubling available")
            if "table_bias" in d:
                flags.append(f"table_bias={d['table_bias']}")
                flags.append(f"sibling_awaiting_deal={d['sibling_awaiting_deal']}")
            flag_str = f" [{', '.join(flags)}]" if flags else ""
            print(f"  {hand_desc} vs {d['dealer_upcard_rank']}{flag_str}: "
                  f"{d['basic_strategy']} → {d['player_action']} "
                  f"({d['samples']} samples, {d['majority']*100:.0f}%)")
        for s in profile["lottery_stakes"]:
            print(f"  dealer lottery, owed={s['owed_bucket']}: "
                  f"avg stake {s['avg_stake']} ({s['samples']} samples)")


if __name__ == "__main__":
    main()
