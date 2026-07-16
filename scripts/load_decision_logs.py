"""
scripts/load_decision_logs.py -- Phase D, step 1 (docs/planning/DecisionLog-Plan.md)

Loads all `decision_log_*.xlsx` exports out of data/decisions/ (the "Hand
Decisions" sheet -- see GET /export_decisions in app/routes/reports.py),
concatenates them, and reports a per-player breakdown -- a first look at how
much training data has been collected for each of Rob, Marco, and David
before any model is attempted.

This is intentionally NOT a training script. It does no feature engineering
or modeling (that's scripts/build_player_profiles.py). It just answers: "how
many decisions do we have per player, and what does the data look like?"

Usage:
    python scripts/load_decision_logs.py
    python scripts/load_decision_logs.py --dir path/to/xlsx
    python scripts/load_decision_logs.py --player Rob
    python scripts/load_decision_logs.py --out data/decisions/combined.csv
"""
from __future__ import annotations

import argparse
import csv
import os
import sys
from collections import Counter, defaultdict
from pathlib import Path

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
DEFAULT_DIR = os.path.join(ROOT, "data", "decisions")


def load_rows(directory: str) -> list[dict]:
    """Read the "Hand Decisions" sheet of every decision_log_*.xlsx in
    `directory` and return all rows as string-keyed dicts (values coerced
    to strings, matching this script's original csv.DictReader shape),
    tagged with the source filename."""
    files = sorted(Path(directory).glob("decision_log_*.xlsx"))
    if not files:
        print(f"No decision_log_*.xlsx files found in {directory}")
        return []

    rows: list[dict] = []
    for path in files:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        if "Hand Decisions" not in wb.sheetnames:
            continue
        ws = wb["Hand Decisions"]
        row_iter = ws.iter_rows(values_only=True)
        header = next(row_iter, None)
        if not header:
            continue
        n = 0
        for values in row_iter:
            row = {str(h): ("" if v is None else str(v)) for h, v in zip(header, values)}
            row["_source_file"] = path.name
            rows.append(row)
            n += 1
        print(f"  loaded {n:4d} rows from {path.name}")
    return rows


def summarize(rows: list[dict], player_filter: str | None = None) -> None:
    if player_filter:
        rows = [r for r in rows if r.get("player", "").lower() == player_filter.lower()]

    if not rows:
        print("No rows to summarize.")
        return

    by_player: dict[str, list[dict]] = defaultdict(list)
    for r in rows:
        by_player[r.get("player", "")].append(r)

    print(f"\nTotal decisions: {len(rows)}")
    print(f"{'Player':<12} {'Total':>6} {'h':>5} {'s':>5} {'d':>5} {'sp':>5} {'ins':>5} {'NPC':>5}")
    for player, prows in sorted(by_player.items(), key=lambda kv: -len(kv[1])):
        actions = Counter(r["action_taken"] for r in prows)
        npc_count = sum(1 for r in prows if r.get("is_npc", "").lower() == "true")
        print(
            f"{player:<12} {len(prows):>6} "
            f"{actions.get('h', 0):>5} {actions.get('s', 0):>5} "
            f"{actions.get('d', 0):>5} {actions.get('sp', 0):>5} "
            f"{actions.get('insurance', 0):>5} {npc_count:>5}"
        )

    # Deviation from basic strategy, per player (human decisions only).
    print("\nDeviation from basic_strategy_action (human decisions, action != insurance):")
    for player, prows in sorted(by_player.items(), key=lambda kv: -len(kv[1])):
        human = [
            r for r in prows
            if r.get("is_npc", "").lower() != "true"
            and r["action_taken"] != "insurance"
            and r.get("basic_strategy_action")
        ]
        if not human:
            continue
        deviations = sum(1 for r in human if r["action_taken"] != r["basic_strategy_action"])
        pct = 100 * deviations / len(human)
        print(f"  {player:<12} {deviations:>4} / {len(human):<4} ({pct:5.1f}%)")

    # Result breakdown (win/loss/push), where backfilled.
    print("\nResult breakdown (where hand_result is filled in):")
    for player, prows in sorted(by_player.items(), key=lambda kv: -len(kv[1])):
        results = Counter(r["hand_result"] for r in prows if r.get("hand_result"))
        if not results:
            continue
        total = sum(results.values())
        print(
            f"  {player:<12} "
            f"win {results.get('win', 0):>3}/{total} "
            f"loss {results.get('loss', 0):>3}/{total} "
            f"push {results.get('push', 0):>3}/{total}"
        )


def write_combined(rows: list[dict], out_path: str) -> None:
    if not rows:
        print("Nothing to write.")
        return
    fieldnames = list(rows[0].keys())
    # _source_file last
    if "_source_file" in fieldnames:
        fieldnames.remove("_source_file")
        fieldnames.append("_source_file")

    out_dir = os.path.dirname(out_path)
    if out_dir:
        os.makedirs(out_dir, exist_ok=True)
    with open(out_path, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow(row)
    print(f"\nWrote {len(rows)} combined rows -> {out_path}")


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--dir", default=DEFAULT_DIR,
        help=f"Directory containing decision_log_*.xlsx files (default: {DEFAULT_DIR})",
    )
    parser.add_argument(
        "--player", default=None,
        help="Only summarize decisions for this player (e.g. Rob)",
    )
    parser.add_argument(
        "--out", default=None,
        help="If set, write all loaded rows to this combined CSV path",
    )
    args = parser.parse_args()

    print(f"Loading decision logs from {args.dir} ...")
    rows = load_rows(args.dir)
    if not rows:
        return 1

    summarize(rows, player_filter=args.player)

    if args.out:
        write_combined(rows, args.out)

    return 0


if __name__ == "__main__":
    sys.exit(main())
