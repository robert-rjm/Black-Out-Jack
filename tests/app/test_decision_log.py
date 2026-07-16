"""
tests/test_decision_log.py
===========================
C5 (Phase C testing) — docs/planning/DecisionLog-Plan.md

Covers:
  1. record_decision captures pre-action state (hand_cards_before /
     hand_total_before) before the action mutates the hand, and
     visible_cards reflects the table at that instant.
  2. The dealer's hidden hole card never appears in visible_cards.
  3. backfill_hand_results fills hand_result for the current round once
     hand.result is set, and leaves other rounds / already-filled rows
     alone.
  4. /export_decisions returns a well-formed XLSX workbook with the
     documented columns across its two sheets, for both Normal mode
     (bet_amount) and Drinking mode (wager).
"""

import io

import pytest

from engine.blackjack import Player, Shoe
from engine.referee import RefereeSession
from app.models.game_room import GameRoom, GameConfig
from app.services.decision_log import (
    record_decision,
    backfill_hand_results,
    _snapshot_visible_cards,
    record_dealer_lottery_entry,
)
from app.services.session_store import game_sessions

from tests.conftest import make_card


# ---------------------------------------------------------------------------
# Helper: build a minimal, started GameRoom
# ---------------------------------------------------------------------------

def _make_room(drinking_mode: bool = True) -> GameRoom:
    rob   = Player("Rob")
    marco = Player("Marco")
    marco.is_dealer = True

    session = RefereeSession([rob, marco], dealer_name="Marco")
    session.start_round(digital=True)
    session.shoe = Shoe(num_decks=4)
    session.shoe.shuffle(quiet=True)

    room = GameRoom(
        session=session,
        config=GameConfig(
            mode="digital",
            drinking_mode=drinking_mode,
        ),
    )
    return room


# ---------------------------------------------------------------------------
# 1. Pre-action snapshot + visible_cards
# ---------------------------------------------------------------------------

def test_record_decision_captures_pre_action_state():
    room = _make_room(drinking_mode=False)
    rob   = room._get_player("Rob")
    marco = room._get_player("Marco")

    # Rob's hand: 10, 6 (=16) -- about to hit
    rob_hand = rob.hands[0]
    rob_hand.cards.extend([make_card("10", "S"), make_card("6", "H")])

    # Dealer shows an upcard + hidden hole card
    marco.dealer_hand.cards.extend([make_card("7", "C"), make_card("K", "D")])

    record_decision(room, rob, rob_hand, "h")

    row = room._decision_log[-1]
    assert row["hand_cards_before"] == "10♠ 6♥"
    assert row["hand_total_before"] == 16
    assert row["is_soft"] is False
    assert row["dealer_upcard"] == "7♣"
    assert row["action_taken"] == "h"
    assert row["round"] == room.round_count
    assert row["player"] == "Rob"
    assert row["hand_index"] == 1

    # Now mutate the hand (simulate the hit) -- the recorded row must NOT change.
    rob_hand.cards.append(make_card("5", "D"))
    assert row["hand_cards_before"] == "10♠ 6♥"
    assert row["hand_total_before"] == 16

    # visible_cards at decision time: Rob's pre-hit hand + dealer upcard only
    # (hole card K♦ excluded).
    visible = row["visible_cards"]
    assert "10♠" in visible and "6♥" in visible
    assert "7♣" in visible
    assert "K♦" not in visible
    assert "5♦" not in visible  # dealt after the decision was recorded


# ---------------------------------------------------------------------------
# 2. Hole card exclusion (direct snapshot check)
# ---------------------------------------------------------------------------

def test_visible_cards_excludes_hidden_dealer_hole_card():
    room = _make_room(drinking_mode=False)
    rob   = room._get_player("Rob")
    marco = room._get_player("Marco")

    rob.hands[0].cards.extend([make_card("9", "S"), make_card("9", "H")])
    marco.dealer_hand.cards.extend([make_card("A", "C"), make_card("Q", "D")])

    visible = _snapshot_visible_cards(room)
    assert "A♣" in visible          # upcard visible
    assert "Q♦" not in visible      # hole card hidden
    assert "9♠" in visible and "9♥" in visible


# ---------------------------------------------------------------------------
# 3. hand_result backfill
# ---------------------------------------------------------------------------

def test_backfill_hand_results():
    room = _make_room(drinking_mode=False)
    rob   = room._get_player("Rob")
    marco = room._get_player("Marco")

    rob.hands[0].cards.extend([make_card("10", "S"), make_card("9", "H")])
    marco.dealer_hand.cards.extend([make_card("7", "C"), make_card("K", "D")])

    record_decision(room, rob, rob.hands[0], "s")
    row = room._decision_log[-1]
    assert row["hand_result"] is None

    # A row from a previous round, already filled -- must stay untouched.
    stale_row = dict(row)
    stale_row["round"] = room.round_count - 1
    stale_row["hand_result"] = "win"
    stale_row["_hand_id"] = -12345
    room._decision_log.append(stale_row)

    # Resolve the round.
    rob.hands[0].result = "win"
    backfill_hand_results(room)

    assert row["hand_result"] == "win"
    assert stale_row["hand_result"] == "win"  # unchanged

    # Calling again is a no-op (idempotent).
    backfill_hand_results(room)
    assert row["hand_result"] == "win"


# ---------------------------------------------------------------------------
# 4. /export_decisions XLSX shape -- one workbook, two sheets
# ---------------------------------------------------------------------------

@pytest.fixture
def app_client():
    from app import create_app
    app = create_app()
    app.testing = True
    return app.test_client()


def _sheet_rows(worksheet):
    """Read an openpyxl worksheet into (header, [row_dict, ...])."""
    it = worksheet.iter_rows(values_only=True)
    header = list(next(it))
    return header, [dict(zip(header, row)) for row in it]


def _populate_and_export(app_client, drinking_mode: bool, room_code: str):
    import openpyxl

    room = _make_room(drinking_mode=drinking_mode)
    rob   = room._get_player("Rob")
    marco = room._get_player("Marco")

    rob.hands[0].cards.extend([make_card("8", "S"), make_card("8", "H")])
    marco.dealer_hand.cards.extend([make_card("6", "C"), make_card("2", "D")])

    record_decision(room, rob, rob.hands[0], "sp")
    room.drinks.last_round_sips["Rob"] = 2
    record_dealer_lottery_entry(room, "Rob", 3, is_npc=False)

    game_sessions[room_code] = room
    try:
        resp = app_client.get(f"/export_decisions?room_code={room_code}")
        assert resp.status_code == 200
        assert resp.mimetype == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

        wb = openpyxl.load_workbook(io.BytesIO(resp.data))
        assert wb.sheetnames == ["Hand Decisions", "Dealer Lottery Entries"]
        return wb
    finally:
        game_sessions.pop(room_code, None)


def test_export_decisions_normal_mode(app_client):
    wb = _populate_and_export(app_client, drinking_mode=False, room_code="TestNormal1")
    header, rows = _sheet_rows(wb["Hand Decisions"])

    assert "_hand_id" not in header
    assert "bet_amount" in header
    assert "wager" in header

    row = rows[0]
    assert row["player"] == "Rob"
    assert row["action_taken"] == "sp"
    assert row["drinking_mode"] is False
    assert row["bet_amount"] is not None   # normal mode -> bet_amount populated
    assert row["wager"] is None            # ... and wager empty


def test_export_decisions_drinking_mode(app_client):
    wb = _populate_and_export(app_client, drinking_mode=True, room_code="TestDrink1")
    _, rows = _sheet_rows(wb["Hand Decisions"])

    row = rows[0]
    assert row["drinking_mode"] is True
    assert row["wager"] is not None        # drinking mode -> wager populated
    assert row["bet_amount"] is None       # ... and bet_amount empty


def test_export_decisions_no_session_returns_404(app_client):
    resp = app_client.get("/export_decisions?room_code=NoSuchRoom")
    assert resp.status_code == 404


# ---------------------------------------------------------------------------
# 5. Dealer Lottery entry capture + export (second sheet of the same workbook)
# ---------------------------------------------------------------------------

def test_record_dealer_lottery_entry_captures_owed_and_context():
    room = _make_room(drinking_mode=True)
    room.drinks.last_round_sips["Rob"] = 3

    record_dealer_lottery_entry(room, "Rob", 4, is_npc=False)

    row = room._dealer_lottery_decision_log[-1]
    assert row["player"] == "Rob"
    assert row["x_entered"] == 4
    assert row["current_owed"] == 3
    assert row["is_npc"] is False
    assert row["num_players"] == 2
    assert row["drinking_mode"] is True
    assert row["round"] == room.round_count


def test_record_dealer_lottery_entry_clamps_negative_owed_to_zero():
    room = _make_room(drinking_mode=True)
    room.drinks.last_round_sips["Rob"] = -2  # bust-vote credit can go negative

    record_dealer_lottery_entry(room, "Rob", 0, is_npc=True)

    assert room._dealer_lottery_decision_log[-1]["current_owed"] == 0


def test_export_decisions_includes_dealer_lottery_sheet(app_client):
    wb = _populate_and_export(app_client, drinking_mode=True, room_code="TestLottery1")
    _, rows = _sheet_rows(wb["Dealer Lottery Entries"])

    row = rows[0]
    assert row["player"] == "Rob"
    assert row["x_entered"] == 3
    assert row["current_owed"] == 2
    assert row["is_npc"] is False
